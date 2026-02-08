import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
import io
import tempfile
import os

# ──────────────────────────────────────────────
#  PAGE CONFIG
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="PDF → Excel Converter",
    page_icon="📊",
    layout="wide",
)

# ──────────────────────────────────────────────
#  CUSTOM CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        background: linear-gradient(90deg, #1F4E79, #2E86C1);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 0;
    }
    .sub-header {
        text-align: center;
        color: #666;
        font-size: 1.1rem;
        margin-bottom: 2rem;
    }
    .stat-card {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 1.2rem;
        border-radius: 12px;
        text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    }
    .stat-number {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1F4E79;
    }
    .stat-label {
        font-size: 0.85rem;
        color: #555;
        margin-top: 4px;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
#  HEADER
# ──────────────────────────────────────────────
st.markdown('<p class="main-header">📊 PDF to Excel Converter</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Upload any bank statement PDF — auto-detects format and exports a beautifully formatted Excel file</p>', unsafe_allow_html=True)

# ──────────────────────────────────────────────
#  DETECTION STRATEGIES
# ──────────────────────────────────────────────

BANK_PROFILES = {
    "icici": {
        "name": "ICICI Bank",
        "keywords": ["ICICI", "icicibank"],
        "skip_keywords": [
            'Detailed', 'Statement', 'Name:', 'Address:', 'A/C No',
            'Jt. Holder', 'Transaction Date from', 'Transaction Period',
            'Statement Request', 'Advanced Search', 'Amount from',
            'Cheque number', 'Transaction remarks', 'Transaction type:',
            'Sl Tran Value', 'No Id Date', 'Page No', 'Closing Balance',
            'Branch:', 'Branch Address', 'A/C Type', 'Cust ID',
            'IFSC Code', 'Account Currency', 'Download', 'from:',
            'WARD', 'PRADESH', 'CAA',
        ],
        "columns": [
            'Sl No', 'Tran Id', 'Value Date', 'Transaction Date',
            'Posted Date', 'Transaction Remarks',
            'Withdrawal (Dr)', 'Deposit (Cr)', 'Balance'
        ],
    },
    "sbi": {
        "name": "SBI",
        "keywords": ["State Bank", "SBI", "sbi.co.in"],
        "skip_keywords": [
            'State Bank', 'Account Statement', 'Account Number',
            'Address', 'Branch', 'IFSC', 'CIF No', 'Nomination',
            'IFS Code', 'MICR Code', 'Page', 'Opening Balance',
        ],
        "columns": [
            'Txn Date', 'Value Date', 'Description', 'Ref No./Cheque No.',
            'Debit', 'Credit', 'Balance'
        ],
    },
    "hdfc": {
        "name": "HDFC Bank",
        "keywords": ["HDFC", "hdfcbank"],
        "skip_keywords": [
            'HDFC BANK', 'Statement of Account', 'Account No',
            'Branch', 'Address', 'IFSC', 'Nomination', 'Page',
            'Opening Balance', 'Closing Balance', 'RTGS', 'NEFT',
        ],
        "columns": [
            'Date', 'Narration', 'Chq./Ref.No.', 'Value Dt',
            'Withdrawal Amt.', 'Deposit Amt.', 'Closing Balance'
        ],
    },
    "generic": {
        "name": "Generic Bank Statement",
        "keywords": [],
        "skip_keywords": [
            'Page', 'Statement', 'Account', 'Branch', 'Address',
            'IFSC', 'Opening Balance', 'Closing Balance',
        ],
        "columns": None,  # auto-detect
    },
}


def detect_bank(text: str) -> str:
    """Detect the bank from the first few pages of text."""
    upper = text.upper()
    for key, profile in BANK_PROFILES.items():
        if key == "generic":
            continue
        for kw in profile["keywords"]:
            if kw.upper() in upper:
                return key
    return "generic"


# ──────────────────────────────────────────────
#  CORE EXTRACTION
# ──────────────────────────────────────────────

def extract_transactions_text(pdf_file, progress_bar=None, status_text=None):
    """
    Extract transactions by parsing text lines from each page.
    Works well for ICICI-style and most digital (non-scanned) bank PDFs.
    """
    pdf = pdfplumber.open(pdf_file)
    total_pages = len(pdf.pages)

    # Detect bank from first 3 pages
    sample_text = ""
    for i in range(min(3, total_pages)):
        t = pdf.pages[i].extract_text()
        if t:
            sample_text += t + "\n"

    bank_key = detect_bank(sample_text)
    profile = BANK_PROFILES[bank_key]

    if status_text:
        status_text.text(f"🏦 Detected: **{profile['name']}** — Processing {total_pages} pages...")

    skip_keywords = profile["skip_keywords"]
    date_pat = re.compile(r'\d{2}/\w{3}/\d{2,4}|\d{2}-\w{3}-\d{2,4}|\d{2}/\d{2}/\d{2,4}|\d{2}-\d{2}-\d{2,4}')
    money_pat = re.compile(r'[\d,]+\.\d{2}')

    all_rows = []

    for i in range(total_pages):
        page = pdf.pages[i]
        text = page.extract_text()
        if not text:
            continue
        lines = text.split('\n')
        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            if any(kw.lower() in stripped.lower() for kw in skip_keywords):
                continue
            # Line must start with a number and contain a date
            match = re.match(r'^(\d+)\s+', stripped)
            if match and date_pat.search(stripped):
                all_rows.append(stripped)

        if progress_bar:
            progress_bar.progress((i + 1) / total_pages)

    pdf.close()
    return all_rows, bank_key, profile


def extract_transactions_table(pdf_file, progress_bar=None, status_text=None):
    """
    Extract using pdfplumber's table extraction — better for some PDFs.
    """
    pdf = pdfplumber.open(pdf_file)
    total_pages = len(pdf.pages)

    sample_text = ""
    for i in range(min(3, total_pages)):
        t = pdf.pages[i].extract_text()
        if t:
            sample_text += t + "\n"

    bank_key = detect_bank(sample_text)
    profile = BANK_PROFILES[bank_key]

    if status_text:
        status_text.text(f"🏦 Detected: **{profile['name']}** — Extracting tables from {total_pages} pages...")

    header_keywords = ["date", "narration", "description", "particular", "debit", "credit",
                        "balance", "withdrawal", "deposit", "amount", "txn", "ref", "cheque",
                        "tran", "value"]
    all_data = []
    headers = None

    for i in range(total_pages):
        page = pdf.pages[i]
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                if not row or not any(cell and cell.strip() for cell in row if cell):
                    continue
                cleaned = [cell.strip() if cell else "" for cell in row]
                row_text = " ".join(cleaned).lower()

                # Detect header row
                if headers is None:
                    matches = sum(1 for kw in header_keywords if kw in row_text)
                    if matches >= 2:
                        headers = cleaned
                        continue

                if headers:
                    all_data.append(cleaned)

        if progress_bar:
            progress_bar.progress((i + 1) / total_pages)

    pdf.close()
    return all_data, headers, bank_key, profile


def parse_text_rows(all_rows, bank_key):
    """Parse raw text rows into structured data."""
    date_pat = re.compile(r'\d{2}/\w{3}/\d{2,4}|\d{2}-\w{3}-\d{2,4}|\d{2}/\d{2}/\d{2,4}|\d{2}-\d{2}-\d{2,4}')
    money_pat = re.compile(r'[\d,]+\.\d{2}')
    parsed = []

    for row_text in all_rows:
        amounts = money_pat.findall(row_text)
        sl_match = re.match(r'^(\d+)\s+', row_text)
        sl_no = sl_match.group(1) if sl_match else ""

        tran_match = re.search(r'(S\d{4,})\s', row_text)
        tran_id = tran_match.group(1) if tran_match else ""

        dates = date_pat.findall(row_text)
        value_date = dates[0] if len(dates) > 0 else ""
        txn_date = dates[1] if len(dates) > 1 else ""

        posted_match = re.search(r'(\d{2}/\d{2}/\d{4})', row_text)
        posted_date = posted_match.group(1) if posted_match else ""

        # Extract remarks
        remainder = row_text
        remainder = re.sub(r'^\d+\s+', '', remainder)
        if tran_id:
            remainder = remainder.replace(tran_id, '', 1)
        for d in dates:
            remainder = remainder.replace(d, '', 1)
        if posted_date:
            remainder = remainder.replace(posted_date, '', 1)
        for a in amounts:
            remainder = remainder.replace(a, '', 1)
        remainder = re.sub(r'\d{2}:\d{2}:\d{2}\s*(AM|PM)', '', remainder)
        remainder = re.sub(r'\s+', ' ', remainder).strip().strip('- ')
        remarks = remainder

        withdrawal = ""
        deposit = ""
        balance = ""

        if len(amounts) >= 3:
            withdrawal = amounts[-3]
            deposit = amounts[-2]
            balance = amounts[-1]
        elif len(amounts) == 2:
            balance = amounts[-1]
            withdrawal = amounts[0]
        elif len(amounts) == 1:
            balance = amounts[0]

        if bank_key == "icici":
            parsed.append({
                'Sl No': sl_no,
                'Tran Id': tran_id,
                'Value Date': value_date,
                'Transaction Date': txn_date,
                'Posted Date': posted_date,
                'Transaction Remarks': remarks,
                'Withdrawal (Dr)': withdrawal,
                'Deposit (Cr)': deposit,
                'Balance': balance,
            })
        else:
            parsed.append({
                'Sl No': sl_no,
                'Date': value_date,
                'Description': remarks,
                'Withdrawal': withdrawal,
                'Deposit': deposit,
                'Balance': balance,
            })

    return pd.DataFrame(parsed)


# ──────────────────────────────────────────────
#  EXCEL FORMATTING
# ──────────────────────────────────────────────

def format_excel(excel_buffer):
    """Apply professional formatting to the Excel workbook."""
    wb = load_workbook(excel_buffer)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    data_font = Font(size=10, name="Calibri")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    debit_font = Font(size=10, name="Calibri", color="CC0000")
    credit_font = Font(size=10, name="Calibri", color="006600")

    # Map column names to indices
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            col_map[val] = col

    # Identify money columns
    money_keywords_dr = ['withdrawal', 'debit', 'dr']
    money_keywords_cr = ['deposit', 'credit', 'cr']
    money_keywords_bal = ['balance']

    dr_cols = [col_map[k] for k in col_map
               if any(m in k.lower() for m in money_keywords_dr)]
    cr_cols = [col_map[k] for k in col_map
               if any(m in k.lower() for m in money_keywords_cr)]
    bal_cols = [col_map[k] for k in col_map
                if any(m in k.lower() for m in money_keywords_bal)]
    all_money_cols = set(dr_cols + cr_cols + bal_cols)

    # Format header
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Format data rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if row % 2 == 0:
                cell.fill = alt_fill

            if col in all_money_cols:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
                if col in dr_cols and cell.value:
                    cell.font = debit_font
                elif col in cr_cols and cell.value:
                    cell.font = credit_font
                else:
                    cell.font = data_font
            else:
                cell.font = data_font

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, 10), 50)

    # Summary rows
    data_last_row = ws.max_row
    sr = data_last_row + 2
    ws.cell(row=sr, column=1, value="SUMMARY").font = Font(
        bold=True, size=12, name="Calibri", color="1F4E79"
    )

    summary_offset = 1
    for col_idx in dr_cols:
        cl = ws.cell(row=1, column=col_idx).column_letter
        label_col = max(1, col_idx - 1)
        ws.cell(row=sr + summary_offset, column=label_col,
                value="Total Withdrawals:").font = Font(bold=True, size=10, name="Calibri")
        c = ws.cell(row=sr + summary_offset, column=col_idx)
        c.value = f"=SUM({cl}2:{cl}{data_last_row})"
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, size=11, color="CC0000", name="Calibri")
        summary_offset += 1

    for col_idx in cr_cols:
        cl = ws.cell(row=1, column=col_idx).column_letter
        label_col = max(1, col_idx - 1)
        ws.cell(row=sr + summary_offset, column=label_col,
                value="Total Deposits:").font = Font(bold=True, size=10, name="Calibri")
        c = ws.cell(row=sr + summary_offset, column=col_idx)
        c.value = f"=SUM({cl}2:{cl}{data_last_row})"
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, size=11, color="006600", name="Calibri")
        summary_offset += 1

    # Freeze header & filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{data_last_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ──────────────────────────────────────────────
#  SIDEBAR CONFIG
# ──────────────────────────────────────────────

with st.sidebar:
    st.header("⚙️ Settings")
    extraction_mode = st.radio(
        "Extraction Method",
        ["Auto (Text-based)", "Table-based"],
        help="**Auto** works best for most bank PDFs. Try **Table-based** if Auto doesn't produce good results."
    )
    st.divider()
    st.markdown("### 🏦 Supported Banks")
    st.markdown("""
    - ✅ ICICI Bank
    - ✅ SBI
    - ✅ HDFC Bank
    - ✅ Any digital PDF with tabular data
    """)
    st.divider()
    st.markdown("### 📋 Output Features")
    st.markdown("""
    - 🎨 Color-coded headers
    - 🔴 Withdrawals in red
    - 🟢 Deposits in green
    - 📊 Alternating row colors
    - 📌 Frozen header row
    - 🔽 Auto-filter enabled
    - ➕ Summary totals
    """)

# ──────────────────────────────────────────────
#  FILE UPLOAD
# ──────────────────────────────────────────────

st.markdown("---")
uploaded_file = st.file_uploader(
    "📂 Upload your Bank Statement PDF",
    type=["pdf"],
    help="Drag and drop or click to browse. Supports large PDFs (100+ pages).",
)

if uploaded_file is not None:
    file_size_mb = uploaded_file.size / (1024 * 1024)
    st.info(f"📄 **{uploaded_file.name}** — {file_size_mb:.1f} MB")

    if st.button("🚀 Convert to Excel", type="primary", use_container_width=True):
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            # Save uploaded file to a temp location
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(uploaded_file.getvalue())
                tmp_path = tmp.name

            status_text.text("📖 Reading PDF...")

            if extraction_mode == "Auto (Text-based)":
                # ── TEXT-BASED EXTRACTION ──
                all_rows, bank_key, profile = extract_transactions_text(
                    tmp_path, progress_bar, status_text
                )
                status_text.text("🔍 Parsing transactions...")
                df = parse_text_rows(all_rows, bank_key)

            else:
                # ── TABLE-BASED EXTRACTION ──
                all_data, headers, bank_key, profile = extract_transactions_table(
                    tmp_path, progress_bar, status_text
                )
                if headers and all_data:
                    max_cols = len(headers)
                    normalized = []
                    for row in all_data:
                        if len(row) < max_cols:
                            row += [""] * (max_cols - len(row))
                        elif len(row) > max_cols:
                            row = row[:max_cols]
                        normalized.append(row)
                    df = pd.DataFrame(normalized, columns=headers)
                else:
                    df = pd.DataFrame()

            # Clean up temp file
            os.unlink(tmp_path)

            if df.empty or len(df) == 0:
                st.error("❌ No transactions found. Try switching the extraction method in the sidebar.")
            else:
                # Convert money columns to numeric
                money_kws = ['withdrawal', 'deposit', 'balance', 'debit', 'credit', 'amount', 'dr', 'cr']
                for col in df.columns:
                    if any(kw in col.lower() for kw in money_kws):
                        df[col] = df[col].apply(
                            lambda x: float(str(x).replace(',', ''))
                            if x and str(x).replace(',', '').replace('.', '').replace('-', '').strip().isdigit() is False
                            and re.match(r'^[\d,]+\.?\d*$', str(x).replace(',', '').strip() or '0')
                            else (float(str(x).replace(',', '')) if x and str(x).replace(',', '').replace('.', '').strip().replace('-', '').isdigit() else None)
                        )

                # Remove fully empty rows
                df = df.dropna(how='all').reset_index(drop=True)

                status_text.text("📊 Generating formatted Excel...")

                # Save to Excel buffer
                excel_buffer = io.BytesIO()
                df.to_excel(excel_buffer, index=False, sheet_name="Bank Statement")
                excel_buffer.seek(0)

                # Apply formatting
                formatted_excel = format_excel(excel_buffer)

                progress_bar.progress(1.0)
                status_text.empty()

                # ── RESULTS ──
                st.success(f"✅ Successfully extracted **{len(df)}** transactions!")

                # Stats cards
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.markdown(f"""<div class="stat-card">
                        <div class="stat-number">{len(df)}</div>
                        <div class="stat-label">Transactions</div>
                    </div>""", unsafe_allow_html=True)
                with col2:
                    st.markdown(f"""<div class="stat-card">
                        <div class="stat-number">{len(df.columns)}</div>
                        <div class="stat-label">Columns</div>
                    </div>""", unsafe_allow_html=True)
                with col3:
                    st.markdown(f"""<div class="stat-card">
                        <div class="stat-number">🏦</div>
                        <div class="stat-label">{profile['name']}</div>
                    </div>""", unsafe_allow_html=True)
                with col4:
                    st.markdown(f"""<div class="stat-card">
                        <div class="stat-number">{file_size_mb:.1f} MB</div>
                        <div class="stat-label">PDF Size</div>
                    </div>""", unsafe_allow_html=True)

                # Download button
                st.markdown("---")
                output_name = uploaded_file.name.replace(".pdf", ".xlsx").replace(".PDF", ".xlsx")
                st.download_button(
                    label="⬇️ Download Excel File",
                    data=formatted_excel,
                    file_name=output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

                # Preview
                st.markdown("### 👀 Data Preview")
                st.dataframe(df.head(50), use_container_width=True, height=400)

                # Column info
                with st.expander("📋 Column Details"):
                    for col in df.columns:
                        non_null = df[col].notna().sum()
                        st.write(f"**{col}** — {non_null} values ({df[col].dtype})")

        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"❌ Error: {str(e)}")
            st.info("💡 **Tips:** Try switching the extraction method in the sidebar, or ensure the PDF is not scanned/image-based.")

else:
    # Empty state
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("### 1️⃣ Upload")
        st.write("Drag & drop or browse for your bank statement PDF")
    with col2:
        st.markdown("### 2️⃣ Convert")
        st.write("Click the convert button — we auto-detect the bank format")
    with col3:
        st.markdown("### 3️⃣ Download")
        st.write("Get a beautifully formatted Excel file with one click")